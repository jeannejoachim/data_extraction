#-----------------------------------------------------------------------------
# Post-treatment script
# See README for more information
# Updated : 01-08-2024
#-----------------------------------------------------------------------------

# Module imports
import os
import sys
import re
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.colors import ListedColormap
import pickle
from datetime import datetime
import logging
import pandas as pd
import openpyxl
import xlsxwriter   # used for Excel conditional formatting
# Function import
from functions import *

####################################################
# USER DEFINED PARAMETERS
####################################################
# TODO define parameters in an external text file after testing on multiple OS
class ParametersFiles:
    # Data folder path, containing data files. The 'r' prefix is important to convert the path in raw string!
    data_folder = r"C:\Users\Jeanne\Documents\Documents_communs\Laura_donnees\Donnees\Raw_data"
    show_plots = False
    # (recommended) if False, plots are only saved in files.
    # if True, plots are shown in the IDE (useful when treating only one sample, or for debugging).
    treat_all_files_in_data_folder = True
    # if False, treatment is performed on given "sample_ID" list (below)
    # if True, "sample_ID" is not used and sample ID is retrieved from all files in data_folder
    sample_ID = ["T1-JAB9780"]
    picture_keyword = ""     # Keyword in filename for test picture
    picture_extension = (".bmp")  # Tuple of strings for possible extension of the test pictures
    # Test result name formatting. Example/ Test result name: E1-JAB9735_indentation
    # with E1-JAB9735 = sample_ID ; E1 = sample name ; indentation = test name
    # => test_name_separator = "_"
    # => sample_ID_position = 0
    # => sample_name_separator = "-"
    # => sample_name_position = 0
    test_name_separator = "_"       # string that separates sample ID to test name in the filename
    sample_ID_position = 0      # 0 = sample ID is the 1st element given (appears before the first file_name_separator)
    sample_name_separator = "-"  # string that separates sample name in sample ID
    sample_name_position = 0
    begin_map = "MAPPING"  # String contained in the line before column name line (entire line will be removed)
    end_map = "Reference"  # String contained in the line after points data (entire line will be removed)
    cropping_frame = 150  # Picture cropping (in pixels) around measurement points limits. 0 = no cropping
    alpha = 0.5  # 0 = transparent, 1 = opaque (used for if plot_thickness_on_image = True)


class ParametersThickness:
    plot_z_curve = True     # Plot position z over time curve for each point for verification purposes
    plot_2D = True          # Plot and save in .png
    plot_on_picture = True  # Plot thickness map on sample real picture
    plot_contour = True  # Plot and save in .png
    plot_3D = True          # Plot and save in .png and .pickle for interactivity
    file_keyword = "thickness"   # Keyword in filename for this test
    file_extension = ".txt"        # File extension for this test
    nb_interp = 100     # Number of points for the data interpolation used to draw color maps
    begin_data = "<DATA>"      # String after which the data point begins
    end_data = "<END DATA>"    # String before which the data point ends


class ParametersIndentation:
    plot_fz_curve_fit = True
    plot_2D = True          # Plot and save in .png
    plot_on_picture = True  # Plot indentation map on sample real picture
    plot_thickness_comparison = True    # Plot indentation vs thickness comparison  map on sample real picture
    nb_interp = 100     # Number of points for the data interpolation used to draw color maps
    fit_start_thickness_percent = 1     # Start value for the Fz fit = thickness percentage (at the same point)
    fit_range_thickness_percent = 12.5    # Range (on thickness percentage) on which Fz fit is performed = linear domain
    # If sensor position values is less than thickness range wanted for fitting, the point is skipped (test failure)
    file_keyword = "indentation"   # Keyword in filename for this test
    file_extension = ".txt"        # File format for this test
    begin_data = "<DATA>"  # String after which the data point begins
    end_data = "<divider>"  # String before which the data point ends
    nu = 0.5        # Poisson coefficient value
    radius = 1   # Sensor radius, in mm
    gravity = 9.80665   # Standard acceleration of gravity, in m/s², used to convert gf/mm² in kPa


####################################################
# INITIALIZING
####################################################
# Pass parameters
prm = ParametersFiles
prm_thickness = ParametersThickness
prm_indentation = ParametersIndentation

# Create logging file
logging.basicConfig(format='%(asctime)s [%(levelname)s] %(message)s', datefmt='%Y/%m/%d %H:%M:%S', encoding='utf-8',
                    handlers=[logging.FileHandler(os.path.join(prm.data_folder, '0_post-treatment-log.log'), mode='w'),
                              logging.StreamHandler(sys.stdout)],
                    level=logging.INFO)
logging.info("================================")
logging.info("Welcome to post-treatment.py")
logging.info("================================")

try:
    ####################################################
    # GET SAMPLES
    ####################################################
    if prm.treat_all_files_in_data_folder:
        all_thickness_files = [f for f in os.listdir(prm.data_folder) if prm_thickness.file_keyword in f
                               if f.endswith(prm_thickness.file_extension)]
        sample_ID = sorted(list(set([f.split(prm.test_name_separator)[prm.sample_ID_position] for f in
                                     all_thickness_files])))
        # set() conversion removes duplicate
        # sorted() sorts the sample names alphabetically, for better Excel sheets management
    else:
        sample_ID = prm.sample_ID

    for s in sample_ID:
        logging.info("[" + str(sample_ID.index(s)+1) + "/" + str(len(sample_ID)) + "] " + "Treating sample " + s)
        # Folder management
        sample_name = s.split(prm.sample_name_separator)[prm.sample_name_position]
        sample_folder = os.path.join(prm.data_folder, sample_name)
        if not os.path.exists(sample_folder):
            os.makedirs(sample_folder)
        raw_extraction_folder = os.path.join(sample_folder, "Raw_extraction")
        if not os.path.exists(raw_extraction_folder):
            os.makedirs(raw_extraction_folder)

        ####################################################
        # THICKNESS
        ####################################################
        logging.info("--------")
        logging.info("1. Thickness data post-treatment:")

        # Extract data
        logging.info("Extracting thickness data...")
        data_points_list_thickness = extract_data(s, prm.data_folder, prm_thickness.file_keyword,
                                                  prm_thickness.file_extension, prm_thickness.begin_data,
                                                  prm_thickness.end_data, tol=1e-2)

        # Compute thickness and interpolate results
        logging.info("Calculating thickness")
        result_thickness = calculate_thickness(data_points_list_thickness)

        ####################################################
        # INDENTATION
        ####################################################
        logging.info("--------")
        logging.info("2. Indentation data post-treatment:")

        # Extract data
        logging.info("Extracting indentation data...")
        data_points_list_indentation = extract_data(s, prm.data_folder, prm_indentation.file_keyword,
                                                    prm_indentation.file_extension, prm_indentation.begin_data,
                                                    prm_indentation.end_data, tol=1e-2)

        # Compute Young modulus from indentation curve fitting
        logging.info("Calculating Young modulus with force data fitting...")

        result_indentation = calculate_young_modulus(data_points_list_indentation, result_thickness['thickness'],
                                                     prm_indentation, prm)

        # If indentation tests has failed, remove corresponding thickness result
        ind_to_remove = np.where(np.isnan(result_indentation['E']))[0]
        result_thickness['thickness'][ind_to_remove] = np.nan

        ####################################################
        # GRAPHS
        ####################################################
        # Plot graphs thickness
        if prm_thickness.plot_z_curve:
            logging.info("Generating z position plots for " + str(len(data_points_list_thickness)) + " points...")
            for i in range(len(data_points_list_thickness)):
                fig = plot_thickness_z_curve(s, i, data_points_list_thickness[i], result_thickness['thickness'][i])
                fig.savefig(os.path.join(raw_extraction_folder, s + "_thickness_z_curve_ID" + str(i+1) + ".png"), bbox_inches='tight')
                if not prm.show_plots:
                    plt.close(fig)
        if prm_thickness.plot_2D:
            logging.info("Generating thickness 2D plot")
            fig = plot_interpolated_surface(s, result_thickness, 'thickness', "[mm]", prm_thickness.nb_interp)
            fig.savefig(os.path.join(sample_folder, s + "_thickness_2D.png"), bbox_inches='tight')
            if not prm.show_plots:
                plt.close(fig)
        if prm_thickness.plot_3D:
            logging.info("Generating thickness 3D plot")
            fig = plot_thickness_3d(s, result_thickness, prm_thickness.nb_interp)
            fig.savefig(os.path.join(sample_folder, s + "_thickness_3D.png"), bbox_inches='tight')
            pickle.dump(fig, open(os.path.join(sample_folder, s + "_thickness_3D.fig.pickle"), 'wb'))
            if not prm.show_plots:
                plt.close(fig)
        if prm_thickness.plot_on_picture:
            logging.info("Generating thickness plot on picture")
            fig = plot_on_picture(s, result_thickness['thickness'], prm, prm_thickness.nb_interp,
                                  colorbar_label="thickness [mm]")
            fig.savefig(os.path.join(sample_folder, s + "_thickness_on_image.png"), bbox_inches='tight')
            if not prm.show_plots:
                plt.close(fig)
        if prm_thickness.plot_contour:
            logging.info("Generating thickness contour plot on picture")
            fig = plot_on_picture(s, result_thickness['thickness'], prm, prm_thickness.nb_interp, contour_plot=True)
            fig.savefig(os.path.join(sample_folder, s + "_thickness_on_image_contour.png"), bbox_inches='tight')
            if not prm.show_plots:
                plt.close(fig)

        # Plot graphs indentation
        if prm_indentation.plot_fz_curve_fit:
            logging.info("Generating Fz curve fit for " + str(len(data_points_list_indentation)) + " points...")
            for i in range(len(data_points_list_indentation)):
                fig = plot_fz_curve_fit(s, i, data_points_list_indentation[i], result_indentation,
                                        prm_indentation.fit_range_thickness_percent)
                fig.savefig(os.path.join(raw_extraction_folder, s + "_fz_curve_fit_ID" + str(i+1) + ".png"), bbox_inches='tight')
                if not prm.show_plots:
                    plt.close(fig)
        if prm_indentation.plot_2D:
            logging.info("Generating Young modulus 2D plot")
            fig = plot_interpolated_surface(s, result_indentation, 'E', "[kPa]", prm_indentation.nb_interp)
            fig.savefig(os.path.join(sample_folder, s + "_young_modulus_2D.png"), bbox_inches='tight')
            if not prm.show_plots:
                plt.close(fig)
        if prm_indentation.plot_on_picture:
            logging.info("Generating Young modulus plot on picture")
            fig = plot_on_picture(s, result_indentation['E'], prm, prm_indentation.nb_interp,
                                  colorbar_label="E [kPa]")
            fig.savefig(os.path.join(sample_folder, s + "_young_modulus_on_image.png"), bbox_inches='tight')
            if not prm.show_plots:
                plt.close(fig)
        if prm_indentation.plot_thickness_comparison:
            logging.info("Generating comparison plot on picture")
            fig = plot_comparison_on_picture(s, result_indentation['E'], result_thickness['thickness'], prm,
                                             prm_indentation.nb_interp, colorbar_label="E [kPa]")
            fig.savefig(os.path.join(sample_folder, s + "_young_modulus_vs_thickness_on_image.png"), bbox_inches='tight')

        ####################################################
        # Excel outputs
        ####################################################
        logging.info("Writing results in Excel file")
        # Store data
        df = pd.DataFrame()
        df["X [mm]"] = result_thickness['pos_x']
        df["Y [mm]"] = result_thickness['pos_y']
        df["Position ID"] = result_thickness['ID']
        df["Type"] = ''
        df["Thickness [mm]"] = result_thickness['thickness']
        df["IM [kPa]"] = result_indentation['E']
        df["R^2"] = result_indentation['corr_coeff']

        # Apply conditional formatting
        cmap = ListedColormap(["steelblue", "lightsteelblue", "white", "lightpink", "lightcoral"])
        df_styled = df.style.background_gradient(axis=None, subset='Thickness [mm]', cmap=cmap)\
            .background_gradient(axis=None, subset='IM [kPa]', cmap=cmap)\
            .apply(highlight_bad_correlation, threshold=0.85, column='R^2', axis=1)

        # Save Sample data in sheet
        excel_path = os.path.join(prm.data_folder, '0_Results.xlsx')
        excel_exists = os.path.exists(excel_path)
        if excel_exists:
            # if file exists: add or rewrite sample sheet
            writer = pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace')
            df_styled.to_excel(writer, sheet_name=s, index=False, na_rep='', freeze_panes=(1, 0))
            writer.close()
            logging.info("Excel file appended:" + excel_path)
        else:
            # if file does not exist: create it
            df_styled.to_excel(excel_path, sheet_name=s, index=False, na_rep='', freeze_panes=(1, 0))
            logging.info("Excel file created")

        # Change column width for "Position ID" and "Thickness [mm]"
        wb = openpyxl.load_workbook(excel_path)
        ws = wb[s]
        for col in [3, 5]:
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

        # Add borders
        thin = openpyxl.styles.borders.Side(border_style="thin", color="000000")
        for row in ws.iter_rows(min_row=2, max_col=np.shape(df)[1], max_row=np.shape(df)[0]+1):
            for cell in row:
                cell.border = openpyxl.styles.borders.Border(top=thin, left=thin, right=thin, bottom=thin)

        # Save file
        wb.save(excel_path)

        ###################################################
        logging.info("Sample treated")
        logging.info("---------------------------------")

    if prm.show_plots:
        plt.show()

except Exception as e:
    logging.critical(str(type(e)) + ": " + str(e))
    raise

logging.info("============== END ==============")
